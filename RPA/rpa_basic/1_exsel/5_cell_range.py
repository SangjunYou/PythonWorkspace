from openpyxl import Workbook
from random import*

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_b = ws["B"] # 영어 column 만 가지고 오기
# for cell in col_b:
#     print(cell.value)

col_range = ws["B:C"]
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]  # 1번째 row만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6]  # 1번째 줄을 제외한 2~6 줄의 내용을 가져옴
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row]   # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ") # A/10 , AZ/250
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")
        print(xy[0], end="")  # A
        print(xy[1], end=" ") # 1
    print()


wb.save("semple.xlsx")
